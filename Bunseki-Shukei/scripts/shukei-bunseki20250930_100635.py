# -*- coding: utf-8 -*-
"""
共通テスト本番レベル模試 集計スクリプト（確定運用版）
- K列ヘッダ「完全修得日」のみで 0/1 判定（状態列は無視）
- cutoff はファイル名から自動（年Y → Y+1の本試験日）。--cutoffで個別上書き可
- 英語/数学/国語の主集計に加え、参考情報（締切超過数の分布）と集計メタデータを出力
"""

import argparse
import re
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


# ====== マスター講座 ======
ENG_TARGETS = [
    "英単語1800", "英熟語750", "英文法750",
    "英例文300", "上級英単語1000", "上級英熟語500",
]
MATH_TARGETS = ["数学Ⅰ", "数学A", "数学Ⅱ", "数学B", "数学ⅠA上級", "数学ⅡB上級"]

# 翌年の本試験日（必要に応じて追記/修正）
EXAM_CUTOFF_BY_YEAR: Dict[int, datetime] = {
    2023: datetime(2023, 1, 17),  # 2022 → 2023/1/17 で締め
    2024: datetime(2024, 1, 17),  # 2023 → 2024/1/17
    2025: datetime(2025, 1, 18),  # 2024 → 2025/1/18
    # 例: 2026: datetime(2026, 1, 17),
}


# ====== ユーティリティ ======
def clean_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().replace("\u00a0", "").replace("\u3000", "") for c in df.columns]
    return df


def normalize_grade_val(x) -> str:
    s = str(x).translate(str.maketrans("０１２３４５６７８９高　 ", "0123456789高  ")).strip()
    return "高3" if re.search(r"(高\s*3|3年|\b3\b)", s) else "高1+2"


def pick_exam_sheets(xls_names):
    """『合同共テ結果』を含むシートのみ対象（例：202501合同共テ結果）"""
    return [s for s in xls_names if "合同共テ結果" in s]


def detect_year_from_stem(stem: str) -> Optional[int]:
    """ファイル名から年を取得。6桁(YYYYMM)優先、なければ4桁(YYYY)。"""
    m6 = re.search(r"(\d{6})", stem)
    if m6:
        return int(m6.group(1)[:4])
    m4 = re.search(r"(\d{4})", stem)
    return int(m4.group(1)) if m4 else None


def cutoff_for_stem(stem: str, override_map: Optional[Dict[str, datetime]] = None) -> Optional[datetime]:
    """
    例: '2024共テ結果' → 翌年 2025 の試験日。
        '202508共テ…' → 2025 → 翌年 2026 の試験日。
    --cutoff で 'YYYYMM:YYYY-MM-DD' を渡した場合はそちらを優先。
    """
    override_map = override_map or {}
    m6 = re.search(r"(\d{6})", stem)
    if m6:
        key6 = m6.group(1)
        if key6 in override_map:
            return override_map[key6]
        year = int(key6[:4])
        return EXAM_CUTOFF_BY_YEAR.get(year + 1)
    year = detect_year_from_stem(stem)
    return EXAM_CUTOFF_BY_YEAR.get(year + 1) if year else None


# ====== 成績（英/数/国）読取 ======
def read_exam(fp: str) -> pd.DataFrame:
    xls = pd.ExcelFile(fp)
    sheets = pick_exam_sheets(xls.sheet_names)
    dfs = []
    for s in sheets:
        df = pd.read_excel(fp, sheet_name=s)
        df = clean_cols(df)

        if "英語" not in df.columns:
            continue

        id_col = next((c for c in df.columns if "生徒番号" in c or c in ["ID", "生徒ID"]), None)
        grade_col = next((c for c in df.columns if "学年" in c), None)
        math1_col = next((c for c in df.columns if "数学Ⅰ・数学A" in c), None)
        math2_col = next((c for c in df.columns if "数学Ⅱ・数学B・数学C" in c), None)
        jpn_col = next((c for c in df.columns if "国語(現古漢)" in c), None)
        if not id_col or not grade_col:
            continue

        sub = pd.DataFrame(
            {
                "student_id": pd.to_numeric(df[id_col], errors="coerce"),
                "grade_norm": df[grade_col].map(normalize_grade_val),
                "score_eng": pd.to_numeric(df["英語"], errors="coerce"),
            }
        )
        if math1_col and math2_col:
            sub["score_math"] = (
                pd.to_numeric(df[math1_col], errors="coerce").fillna(0)
                + pd.to_numeric(df[math2_col], errors="coerce").fillna(0)
            )
        else:
            sub["score_math"] = pd.NA
        sub["score_jpn"] = pd.to_numeric(df[jpn_col], errors="coerce") if jpn_col else pd.NA

        sub = sub.dropna(subset=["student_id"]).copy()
        sub["student_id"] = sub["student_id"].astype(int).astype(str)
        dfs.append(sub[["student_id", "grade_norm", "score_eng", "score_math", "score_jpn"]])

    if dfs:
        out = pd.concat(dfs, ignore_index=True)
        out = out.groupby(["student_id", "grade_norm"], as_index=False).agg(
            score_eng=("score_eng", "max"),
            score_math=("score_math", "max"),
            score_jpn=("score_jpn", "max"),
        )
    else:
        out = pd.DataFrame(columns=["student_id", "grade_norm", "score_eng", "score_math", "score_jpn"])
    return out


# ====== 講座シート名（厳格：YYYY + 講座名、末尾一致） ======
def _sheets_for_targets(sheet_names, target_list):
    def norm(x: str) -> str:
        return re.sub(r"\s+", "", str(x))

    compiled = []
    for t in target_list:
        # ^\s*\d{4}\s*.*英単語1800\s*$（空白ゆれ無視、末尾一致）
        pat = re.compile(rf"^\s*\d{{4}}\s*.*{re.escape(norm(t))}\s*$", re.I)
        compiled.append((t, pat))

    hit = []
    for s in sheet_names:
        s_norm = norm(s)
        for t, pat in compiled:
            if pat.match(s_norm):
                hit.append((t, s))  # (講座名, 実シート名)
                break
    return hit


# ====== 高速基礎マスター：K列（完全修得日）のみで判定 ======
def _read_kdate_series(df: pd.DataFrame) -> Optional[pd.Series]:
    """K列(=index10)または列名一致で『完全修得日』を返す。無ければNone。"""
    if "完全修得日" in df.columns:
        return pd.to_datetime(df["完全修得日"], errors="coerce")
    if df.shape[1] > 10 and str(df.columns[10]).strip() == "完全修得日":
        return pd.to_datetime(df.iloc[:, 10], errors="coerce")
    return None


def count_mastery(fp: str, target_list: list, cutoff: Optional[datetime]) -> pd.DataFrame:
    xls = pd.ExcelFile(fp)
    hit_pairs = _sheets_for_targets(xls.sheet_names, target_list)
    per_target = []

    for tgt in target_list:
        frames = []
        for t, s in hit_pairs:
            if t != tgt:
                continue
            df = pd.read_excel(fp, sheet_name=s)
            df = clean_cols(df)

            id_col = next((c for c in df.columns if "生徒番号" in c or c in ["ID", "生徒ID"]), None)
            if not id_col:
                continue
            dates = _read_kdate_series(df)
            if dates is None:
                continue

            sub = pd.DataFrame({
                "student_id": pd.to_numeric(df[id_col], errors="coerce"),
                "completed": dates.notna() & ((dates <= cutoff) if cutoff is not None else dates.notna())
            }).dropna(subset=["student_id"])
            sub["student_id"] = sub["student_id"].astype(int).astype(str)
            sub["completed"] = sub["completed"].astype(int)
            sub = sub.groupby("student_id", as_index=False)["completed"].max()  # 同講座の重複は 0/1 の max
            frames.append(sub[["student_id", "completed"]])

        if frames:
            tgt_df = (
                pd.concat(frames, ignore_index=True)
                .groupby("student_id", as_index=False)["completed"].max()
            )
        else:
            tgt_df = pd.DataFrame(columns=["student_id", "completed"])

        per_target.append(tgt_df.rename(columns={"completed": tgt}))

    if not per_target:
        return pd.DataFrame(columns=["student_id", "mas_count"])

    base = per_target[0]
    for add in per_target[1:]:
        base = base.merge(add, on="student_id", how="outer")

    for tgt in target_list:
        if tgt not in base.columns:
            base[tgt] = 0
        base[tgt] = base[tgt].fillna(0).astype(int)

    base["mas_count"] = base[target_list].sum(axis=1).astype(int)
    return base[["student_id", "mas_count"]]


# ====== 参考：締切超過数（cutoffより後に達成した講座の個数） ======
def count_late_mastery(fp: str, target_list: list[str], cutoff: Optional[datetime]) -> pd.DataFrame:
    if cutoff is None:
        return pd.DataFrame(columns=["student_id", "late_count"])

    xls = pd.ExcelFile(fp)
    hit_pairs = _sheets_for_targets(xls.sheet_names, target_list)
    per_target = []

    for tgt in target_list:
        frames = []
        for t, s in hit_pairs:
            if t != tgt:
                continue
            df = pd.read_excel(fp, sheet_name=s)
            df = clean_cols(df)
            id_col = next((c for c in df.columns if "生徒番号" in c or c in ["ID", "生徒ID"]), None)
            if not id_col:
                continue
            dates = _read_kdate_series(df)
            if dates is None:
                continue

            sub = pd.DataFrame({
                "student_id": pd.to_numeric(df[id_col], errors="coerce"),
                "date": dates
            }).dropna(subset=["student_id"])
            sub["student_id"] = sub["student_id"].astype(int).astype(str)
            sub = sub.groupby("student_id", as_index=False)["date"].min()  # 最初の達成日で評価
            frames.append(sub)

        if frames:
            tgt_df = pd.concat(frames, ignore_index=True)
        else:
            tgt_df = pd.DataFrame(columns=["student_id", "date"])

        tgt_df[tgt] = (tgt_df["date"] > cutoff).fillna(False).astype(int)
        per_target.append(tgt_df[["student_id", tgt]])

    if not per_target:
        return pd.DataFrame(columns=["student_id", "late_count"])

    base = per_target[0]
    for add in per_target[1:]:
        base = base.merge(add, on="student_id", how="outer")

    for tgt in target_list:
        if tgt not in base.columns:
            base[tgt] = 0
        base[tgt] = base[tgt].fillna(0).astype(int)

    base["late_count"] = base[target_list].sum(axis=1).astype(int)
    return base[["student_id", "late_count"]]


# ====== 表作成 ======
def summarize_long(df, key, score_col, label, max_val):
    template = pd.MultiIndex.from_product(
        [["高1+2", "高3"], list(range(0, max_val + 1))], names=["学年区分", key]
    ).to_frame(index=False)
    g = (
        df.dropna(subset=[score_col])
        .groupby(["grade_norm", key])
        .agg(人数=("student_id", "nunique"), 平均得点=(score_col, "mean"))
        .reset_index()
        .rename(columns={"grade_norm": "学年区分"})
    )
    out = template.merge(g, on=["学年区分", key], how="left")
    out["人数"] = out["人数"].fillna(0).astype(int)
    out[label] = out["平均得点"].round(1)
    return out.drop(columns=["平均得点"]).sort_values(["学年区分", key]).reset_index(drop=True)


def summarize_wide(long_df, key_col, value_col, label_left="完全修得数"):
    df = long_df[["学年区分", key_col, value_col]].copy()
    wide = df.pivot_table(index=key_col, columns="学年区分", values=value_col, aggfunc="first").reset_index()
    wide = wide.rename(columns={key_col: label_left})
    cols = [label_left] + [c for c in ["高1+2", "高3"] if c in wide.columns]
    return wide[cols]


def summarize_distribution(df: pd.DataFrame, key: str, max_val: int) -> pd.DataFrame:
    """学年×(0..max) の分布（人数のみ）"""
    template = pd.MultiIndex.from_product(
        [["高1+2", "高3"], list(range(0, max_val + 1))], names=["学年区分", key]
    ).to_frame(index=False)
    g = (
        df.groupby(["grade_norm", key])
          .agg(人数=("student_id", "nunique"))
          .reset_index()
          .rename(columns={"grade_norm": "学年区分"})
    )
    out = template.merge(g, on=["学年区分", key], how="left")
    out["人数"] = out["人数"].fillna(0).astype(int)
    return out.sort_values(["学年区分", key]).reset_index(drop=True)


def autosize_and_style(ws, header_row_idx: int):
    from openpyxl.utils import get_column_letter
    dims = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value)))
    for cidx, w in dims.items():
        ws.column_dimensions[get_column_letter(cidx)].width = min(w + 2, 28)
    for c in ws[header_row_idx]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"


def save_sheet(wb, title, df):
    ws = wb.create_sheet(title)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    autosize_and_style(ws, 1)


# ====== メイン ======
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", action="append", required=True, help="対象Excel（複数可）")
    ap.add_argument("--cutoff", action="append", default=[], help="上書き: YYYYMM:YYYY-MM-DD（複数可）")
    ap.add_argument("--out", default="集計結果.xlsx", help="出力Excelファイル名")
    args = ap.parse_args()

    # 個別上書き（YYYYMM単位）
    cutoff_override: Dict[str, datetime] = {}
    for pair in args.cutoff:
        key, datestr = pair.split(":", 1)
        cutoff_override[key] = datetime.fromisoformat(datestr)

    records = []
    meta_rows = []
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for fp in args.file:
        stem = Path(fp).stem
        cutoff = cutoff_for_stem(stem, cutoff_override)

        base = read_exam(fp)
        eng_m = count_mastery(fp, ENG_TARGETS, cutoff).rename(columns={"mas_count": "eng_mas"})
        math_m = count_mastery(fp, MATH_TARGETS, cutoff).rename(columns={"mas_count": "math_mas"})
        merged = base.merge(eng_m, on="student_id", how="left").merge(math_m, on="student_id", how="left")
        merged["eng_mas"] = merged["eng_mas"].fillna(0).astype(int)
        merged["math_mas"] = merged["math_mas"].fillna(0).astype(int)
        merged["tot_mas"] = merged["eng_mas"] + merged["math_mas"]
        merged["yyyymm_or_year"] = stem

        # 参考：締切超過数（英/数）
        eng_late = count_late_mastery(fp, ENG_TARGETS, cutoff).rename(columns={"late_count": "eng_late"})
        math_late = count_late_mastery(fp, MATH_TARGETS, cutoff).rename(columns={"late_count": "math_late"})
        merged = merged.merge(eng_late, on="student_id", how="left").merge(math_late, on="student_id", how="left")
        merged["eng_late"] = merged["eng_late"].fillna(0).astype(int)
        merged["math_late"] = merged["math_late"].fillna(0).astype(int)

        records.append(merged)

        # メタデータ
        meta_rows.append({
            "ファイル": Path(fp).name,
            "キー(年/年月)": stem,
            "cutoff(適用日)": cutoff.strftime("%Y-%m-%d") if cutoff else "",
            "生徒数(一意ID)": merged["student_id"].nunique(),
        })

    if not records:
        print("⚠ データがありません。--file を指定してください。")
        return

    panel = pd.concat(records, ignore_index=True)

    # 主集計（長い表）
    t_eng = summarize_long(panel, "eng_mas", "score_eng", "平均英語得点", 6)
    t_math = summarize_long(panel, "math_mas", "score_math", "平均数学得点", 6)
    t_jpn = summarize_long(panel, "tot_mas", "score_jpn", "平均国語得点", 12)

    # 棒グラフ用（横持ち）
    w_eng = summarize_wide(t_eng, "eng_mas", "平均英語得点", label_left="英語 完全修得数")
    w_math = summarize_wide(t_math, "math_mas", "平均数学得点", label_left="数学 完全修得数")
    w_jpn = summarize_wide(t_jpn, "tot_mas", "平均国語得点", label_left="英数合計 完全修得数")

    # 参考：締切超過数の分布
    ref_eng = summarize_distribution(panel, "eng_late", 6)
    ref_math = summarize_distribution(panel, "math_late", 6)

    # メタデータ
    meta_df = pd.DataFrame(meta_rows)
    meta_df.insert(0, "実行時刻", run_ts)

    # Excel出力
    wb = Workbook()
    ws = wb.active
    ws.title = "英語マスター×平均英語（長）"
    for r in dataframe_to_rows(t_eng, index=False, header=True):
        ws.append(r)
    autosize_and_style(ws, 1)

    save_sheet(wb, "数学マスター×平均数学（長）", t_math)
    save_sheet(wb, "英数合計×平均国語（長）", t_jpn)
    save_sheet(wb, "棒グラフ用_英語（横）", w_eng)
    save_sheet(wb, "棒グラフ用_数学（横）", w_math)
    save_sheet(wb, "棒グラフ用_国語（横）", w_jpn)
    save_sheet(wb, "参考_英語 締切超過数の分布", ref_eng)
    save_sheet(wb, "参考_数学 締切超過数の分布", ref_math)
    save_sheet(wb, "集計メタデータ", meta_df)

    wb.save(args.out)
    print("✅ 出力完了：", args.out)


if __name__ == "__main__":
    main()
